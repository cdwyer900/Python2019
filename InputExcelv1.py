#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sat May 12 11:06:36 2018

@author: j2

read in from a excel file into a list

first column - id num can discard
second column - date plus country of origin in a string - keep
third column - url keep plus name of publication
fourth - discard A
fifth - discard B
sixth - discard C
seventh - discard

files are CAData1.xlsx
CAData2.xlsx
CAData3.xlsx

"""
import xlrd # for xls files
from pandas import DataFrame
import re

import openpyxl # for xlsx files
from openpyxl import load_workbook

def wb_to_txt_file(in_wb, out_file):
    # print(wb1.sheetnames)
    sheets = []
    sheets = in_wb.sheetnames

    line = ""
    old_line = "" # get rid of duplicates
    date = ""
    country = ""

    # for each sheet
    for i in range(len(sheets)):
        s = in_wb[sheets[i]]
        # for each sheet skip first four rows
        for row in range(5,(s.max_row+1)):
            if len(s.cell(row=row,column=7).value)<50:
                break
            l1 = (s.cell(row=row,column=2).value).split()
            # print(l1)
            date = l1[0]
            # print(date)
            country = l1[1]
            l2 = s.cell(row=row,column=3).value #publisher
           
            l3 = s.cell(row=row,column=3).hyperlink.target #url
            line = date+', '+country+', '+l2+', '+l3+'\n'
            #  print (line)
            if line != old_line:
                out_file.write(line)
            old_line = line
        
    
#main 
wb1 = load_workbook('CAData1.xlsx')
wb2 = load_workbook('CAData2.xlsx')
wb3 = load_workbook('CAData3.xlsx')

fout = open('CADataOut.txt','at')
fout.truncate() # clear out file

wb_to_txt_file(wb1, fout)
wb_to_txt_file(wb2, fout)
wb_to_txt_file(wb3, fout)

fout.close()
    
    
        








